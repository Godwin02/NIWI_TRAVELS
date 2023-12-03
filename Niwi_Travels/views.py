from itertools import count
from django.forms import ValidationError
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import *
from django.core.paginator import Paginator
from django.core.cache import cache
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.sessions.models import Session

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import User, Traveller, Driver
from django.core.paginator import Paginator
from django.core.cache import cache
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.debug import sensitive_post_parameters
from django.views.decorators.cache import never_cache

# Import User model dynamically to ensure compatibility
User = get_user_model()

def register(request):
    if request.user.is_authenticated:
        return redirect('/')
    return render(request, 'register.html')

def index(request):
    running_packages = TravelPackage.objects.filter(status='Running')

    return render(request, 'index.html',{'running_packages': running_packages})

# Define a view to activate the user's email
def activate_email(request, uidb64, token):
    try:
        # Decode the uid and get the user
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
        
        # Verify the token
        if default_token_generator.check_token(user, token):
            # Activate the user's account
            user.is_active = True
            user.save()
            messages.success(request, "Your account has been activated. You can now log in.")
            return redirect('/log')
        else:
            messages.error(request, "Invalid activation link.")
            return redirect('/log')
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
        messages.error(request, "Invalid activation link.")
        return redirect('/log')
    
@csrf_protect
@sensitive_post_parameters()
@never_cache
@csrf_exempt
def drireg(request):
    if request.user.is_authenticated:
        return redirect('/dhome')

    if request.method == 'POST':
        # username = request.POST['username']
        username = request.POST['email']
        password = request.POST['password']

        # Check if an active user with the same username exists
        if User.objects.filter(email=username, is_active=True,is_driver=True).exists():
            messages.error(request, " Email is already taken.")
            return redirect('/dregister')
        cache.clear()
        # Create a user with is_active=False initially
        user = User(username=username,email=username, is_driver=True ,is_traveller=False, is_active=False)
        user.set_password(password)

        try:
            user.save()
            driver = Driver(user=user)
            driver.save()
        except ValidationError as e:
            # Handle validation errors, e.g., unique constraint violations
            messages.error(request, e.messages[0])  # Display the first error message
            return redirect('dregister')

        # Generate an activation token for the user
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        # Create an activation link with the token
        activation_link = reverse('activate_email', args=[uid, token])

        # Construct the activation email
        subject = 'Activate Your Account'
        message = render_to_string('activation_email.html', {
            'user': user,
            'domain': get_current_site(request).domain,
            'activation_link': activation_link,
        })
        from_email = 'godwinbmenachery2024a@mca.ajce.in'  # Change this to your email
        to_email = user.email

        # Send the activation email
        send_mail(subject, message, from_email, [to_email])
        messages.success(request, "A link has been sent to your registered email. Please click on it to continue.")
        return render(request, 'index.html', {'email': user.email})

    cache.clear()
    return render(request, 'dregister.html')

def login(request):
    if request.user.is_authenticated and request.user.is_traveller:
        return redirect('/thome')
    if request.user.is_authenticated and request.user.is_driver:
        return redirect('/dhome')
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('/admins')

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)

        if user is not None:
            auth_login(request, user)
            
            if user.is_traveller==True:
                request.session["user_type"] = "traveller" 
                print("hello") # Set session variable for user type
                messages.success(request, "welcome To NIWI TRAVELS. Explore the Unexplored!")
                return redirect('/thome')
            elif user.is_driver==True:
                request.session["user_type"] = "driver"  # Set session variable for user type
                return redirect('/dhome')
            elif user.is_staff:
                request.session["user_type"] = "admin"  # Set session variable for user type
                return redirect('/admins')
        else:
            messages.error(request, "Username or Password is Incorrect.")
    
    cache.clear()
    return render(request, 'login.html')

@csrf_protect
@sensitive_post_parameters()
@never_cache
@csrf_exempt
def tregister(request):
    if request.user.is_authenticated:
        return redirect('/thome')

    if request.method == 'POST':
        # username = request.POST['username']
        username = request.POST['email']
        password = request.POST['password']
        cache.clear()
        # Check if an active user with the same email exists
        if User.objects.filter(email=username, is_active=True,is_traveller=True).exists():
            messages.error(request, "An account with this email already exists.")
            return redirect('tregister')
        cache.clear()

        # Create a user with is_active=False initially
        user = User(username=username, email=username, is_traveller=True, is_driver=False,is_active=False)
        user.set_password(password)

        try:
            user.full_clean()  # Validate the user data
            user.save()
            traveller = Traveller(user=user)
            traveller.save()
        except ValidationError as e:
            # Handle validation errors, e.g., unique constraint violations
            messages.error(request, e.messages[0])  # Display the first error message
            return redirect('tregister')

        # Generate an activation token for the user
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        # Create an activation link with the token
        activation_link = reverse('activate_email', args=[uid, token])

        # Construct the activation email
        subject = 'Activate Your Account'
        message = render_to_string('activation_email.html', {
            'user': user,
            'domain': get_current_site(request).domain,
            'activation_link': activation_link,
        })
        from_email = 'godwinbmenachery2024a@mca.ajce.in'  # Change this to your email
        to_email = user.email

        # Send the activation email
        send_mail(subject, message, from_email, [to_email])
        messages.success(request, "A link has been sent to your registered email. Please click on it to continue.")

        return render(request, 'index.html', {'email': user.email})

    cache.clear()
    return render(request, 'tregister.html')


def activate_email(request, uidb64, token):
    try:
        # Decode the uid and get the user
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
        
        # Verify the token
        if default_token_generator.check_token(user, token):
            # Activate the user's account
            user.is_active = True
            user.save()
            messages.success(request, "Your account has been activated. You can now log in.")
            return redirect('log')
        else:
            messages.error(request, "Invalid activation link.")
            return redirect('log')
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
        messages.error(request, "Invalid activation link.")
        return redirect('log')

@never_cache
@login_required(login_url='log')
def traveller_home(request):
    current_date = timezone.now().date()
    tour=TravelPackage.objects.all()
    for packages in tour:
        if current_date > packages.start_date:
            packages.status = 'Pending'
            packages.save()


    running_packages = TravelPackage.objects.filter(
        status='Running',
        start_date__gt=current_date,feed='Post'
    )
    if request.user.is_authenticated and hasattr(request.user, 'traveller'):
        profile = request.user.traveller  # Get or create the Traveller model instance
        return render(request, 'loginview.html', {'profile': profile, 'running_packages': running_packages})
    if request.user.is_traveller:
        profile = request.user.is_traveller  # Corrected from 'is_traveller' to 'traveller'
        return render(request, 'loginview.html', {'profile': profile, 'running_packages': running_packages})
    # If the user is not a traveler or not logged in, redirect to login
    return render(request, 'login.html')

@never_cache
@login_required(login_url='log')
def update_traveler(request):
    if request.method == 'POST':
        user = request.user
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        phone_number = request.POST['phone_number']
        gender = request.POST['gender']
        country = request.POST['country']
        profile_photo = request.FILES.get('profile_photo')

        traveler, created = Traveller.objects.get_or_create(user=user)
        traveler.first_name = first_name
        traveler.last_name = last_name
        traveler.phone_number = phone_number
        traveler.gender = gender
        traveler.country = country
        if profile_photo:
            traveler.profile_photo = profile_photo
        traveller = Traveller(user=user, first_name=first_name, last_name=last_name, phone_number=phone_number, country=country, gender=gender, profile_photo=profile_photo)
        traveller.save()
        traveler.save()
        messages.success(request, "Profile updated successfully.")
        return redirect('/profile_updated')
    
    # Moved user_details definition out of the conditional block
    user_details = {
        'username': request.user.username,
        'email': request.user.email,
    }
    if request.user.is_authenticated and hasattr(request.user, 'traveller'):
        user = request.user
        profile = request.user.traveller # Get or create the Traveller model instance

        return render(request, 'update_traveler.html', {'profile': profile, 'user_details': user_details})
    if request.user.is_traveller:
        profile = request.user.is_traveller  # Corrected from 'is_traveller' to 'traveller'

        return render(request, 'update_traveler.html', {'profile': profile, 'user_details': user_details})
    return render(request, 'login.html')  # Handle the case when the user is not authenticated or not a traveler

    

@login_required(login_url='log')
def profile_updated(request):
    return render(request, 'tprofile_updated.html')

@never_cache
@login_required(login_url='log')
def viewprofile(request):
    if request.user.is_authenticated and hasattr(request.user, 'traveller'):
        profile = request.user.traveller # Get or create the Traveller model instance

        # Get user details from the Traveller model
        user_details = {
            'username': request.user.username,
            'email': request.user.email,
            }
        

        return render(request, 'viewprofileT.html', {'profile': profile, 'user_details': user_details})
    if request.user.is_traveller:
                profile = request.user.is_traveller  # Retrieve the traveler's profil # Replace 'update_traveller' with the actual URL name for your update page
                user_details = {
                    'username': request.user.username,
                    'email': request.user.email,
                    # 'first_name': request.user.first_name,
                    # 'last_name': request.user.last_name,
                    # 'phone_number': request.user.phone_number,
                    # 'country': request.user.country,
                    # 'gender': request.user.gender,
                    # 'profile_photo': Traveller.profile_photo if Traveller.profile_photo else None,
                }
                return render(request, 'viewprofileT.html', {'profile': profile, 'user_details': user_details})
    return render(request, 'login.html')  # Handle the case when the user is not authenticated or not a traveler



@login_required(login_url='log')
def update_driver(request):
    if request.method == 'POST':
        user = request.user
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        contact_email = request.POST['contact_email']
        contact_phone_number = request.POST['contact_phone_number']
        date_of_birth = request.POST['date_of_birth']
        location = request.POST['location']
        profile_photo = request.FILES.get('profile_photo')
        license = request.FILES.get('license')

        driver, created = Driver.objects.get_or_create(user=user)
        driver.first_name = first_name
        driver.last_name = last_name
        driver.contact_email = contact_email
        driver.contact_phone_number = contact_phone_number
        driver.date_of_birth = date_of_birth
        driver.location = location

        if profile_photo:
            driver.profile_photo = profile_photo
        
        if license:
            driver.license = license

        driver.save()
        messages.success(request, "Profile updated successfully.")
        return redirect('/dpro_updated')
    
    cache.clear()
    if request.user.is_driver:
        profile = request.user.driver
        user_details = {
            'username': request.user.username,
            'email': request.user.email,
        }
    
    return render(request, 'update_driver.html', {'profile': profile, 'user_details': user_details})

@login_required(login_url='log')
def dprofile_updated(request):
    return render(request, 'dprofile_updated.html')

@never_cache
@login_required(login_url='log')
def driver_home(request):
    user_type = request.session.get("user_type")  # Get user type from session
    
    if user_type == "driver":
        profile = request.user.driver
        return render(request, 'driver.html', {'profile': profile})
    
    # If user is not a driver or not logged in, redirect to login
    return redirect('/log')

@login_required(login_url='log')
def viewprofileD(request):
    user_type = request.session.get("user_type")  # Get user type from session
    
    if user_type == "driver":
        profile = request.user.driver
        user_details = {
            'username': request.user.username,
            'email': request.user.email,
        }
        return render(request, 'viewprofileD.html', {'profile': profile, 'user_details': user_details})
    
    # If user is not a driver or not logged in, redirect to login
    return redirect('/log')

def logo(request):
    logout(request)
    cache.clear()
    return redirect('/')


@never_cache
@login_required(login_url='log')
def admin(request):
    if request.user.is_staff:
        return render(request, 'adminreg.html')
    cache.clear()
    return redirect('/log')


@never_cache
@login_required(login_url='log')
def user_list(request):
    if request.user.is_staff:
        user_profiles = User.objects.all()
        user_count = user_profiles.count()
        paginator = Paginator(user_profiles, 10)
    
        page_number = request.GET.get('page')
        page = paginator.get_page(page_number)
        context = {'user_profiles': page, 'user_count': user_count}
        return render(request, 'user_list.html', context)
    cache.clear()
    return redirect('/log')

@never_cache
@login_required(login_url='log')
def travellers(request):
    if request.user.is_staff:
        user_profiles = User.objects.filter(is_traveller=True)
        user_count = user_profiles.count()
        paginator = Paginator(user_profiles, 10)
    
        page_number = request.GET.get('page')
        page = paginator.get_page(page_number)
        context = {'user_profiles': page, 'user_count': user_count}
        return render(request, 'travellers.html', context)
    return redirect('/log')


@never_cache
@login_required(login_url='log')
def drivers(request):
    if request.user.is_staff:
        user_profiles = User.objects.filter(is_driver=True)
        user_count = user_profiles.count()
        paginator = Paginator(user_profiles, 10)
    
        page_number = request.GET.get('page')
        page = paginator.get_page(page_number)
        
        approved_users = User.objects.filter(driver__verification='approved')
        rejected_users = User.objects.filter(driver__verification='rejected')
        approved_drivers = Driver.objects.filter(verification='approved')
        rejected_drivers = Driver.objects.filter(verification='rejected')

        context = {
            'user_profiles': page,
            'user_count': user_count,
            'approved_users': approved_users,
            'rejected_users': rejected_users,
            'approved_drivers': approved_drivers,
            'rejected_drivers': rejected_drivers,
        }

        return render(request, 'drivers.html', context)
    return redirect('/log')



from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags

@never_cache
@login_required(login_url='log')
def update_verification_status(request, user_profile_id):
    if request.method == 'POST':
        user_profile = User.objects.get(pk=user_profile_id)
        new_status = request.POST.get('verification_status')

        if new_status in ('approved', 'rejected'):
            old_status = user_profile.driver.verification

            user_profile.driver.verification = new_status
            user_profile.driver.save()

            # Check if the verification status has changed
            if old_status != new_status:
                # Send an email notification
                subject = 'Verification Status Changed'
                message = f'Your verification status has been changed to "{new_status}".'
                from_email = 'godwinbmenachery2024a@mca.ajce.in'
                recipient_list = [user_profile.email]

                send_mail(subject, message, from_email, recipient_list, fail_silently=False)
    
    return redirect('/drivers')  # Redirect back to the user profile list view



@never_cache
@login_required(login_url='log')
def list_approved_users(request):
    # Query the User and Driver models to get approved users
    approved_users = User.objects.filter(driver__verification='approved')
    approved_drivers = Driver.objects.filter(verification='approved')

    context = {'approved_users': approved_users, 'approved_drivers': approved_drivers}

    return render(request, 'drivers.html', context)


from django.contrib.auth import views as auth_views
from django.urls import reverse_lazy

# Use the built-in views provided by Django's auth app
class MyPasswordResetView(auth_views.PasswordResetView):
    template_name = 'password_reset_form.html'  # Your template for the password reset form
    email_template_name = 'password_reset_email.html'  # Your email template for the password reset email
    success_url = reverse_lazy('password_reset_done')  # URL to redirect after successful form submission

class MyPasswordResetDoneView(auth_views.PasswordResetDoneView):
        template_name = 'password_reset_done.html'  # Your template for password reset done page


class MyPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'password_reset_confirm.html'  # Your template for password reset confirmation form
    success_url = reverse_lazy('password_reset_complete')  # URL to redirect after successful password reset

class MyPasswordResetCompleteView(auth_views.PasswordResetCompleteView):
    template_name = 'password_reset_complete.html'  # Your template for password reset complete page

    def get(self, request, *args, **kwargs):
        return redirect('/log')  # Replace 'home' with the desired URL name




# views.py

from django.shortcuts import render, redirect
from django.http import HttpResponse
# from .models import TravelPackage  # Import your TravelPackage model
from django.utils.datastructures import MultiValueDictKeyError




from .models import TravelPackage
from .models import PackageImage  # Import the PackageImage model

@never_cache
@login_required(login_url='log')
def upload_package(request):
    if request.method == "POST":
        # This block handles form submission and package creation when the form is submitted.

        # Retrieve form data from the request
        package_name = request.POST.get("package_name")
        description = request.POST.get("description")
        destination = request.POST.get("destination")
        duration = request.POST.get("duration")
        price = request.POST.get("price")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        accommodation = request.POST.get("accommodation")
        # meals = request.POST.get("meals")
        transportation = request.POST.get("transportation")
        activities = request.POST.get("activities")

        selected_inclusions = [request.POST.get(key) for key in request.POST if key.startswith("inclusions")]
        inclusions = ', '.join(selected_inclusions)

        # Process and build a comma-separated string for exclusions
        selected_exclusions = [request.POST.get(key) for key in request.POST if key.startswith("exclusions")]
        exclusions = ', '.join(selected_exclusions)

        images = request.FILES.getlist("images")  # Handle multiple uploaded images
        # ratings = request.POST.get("ratings")
        availability = request.POST.get("availability")
        booking_deadline = request.POST.get("booking_deadline")
        category = request.POST.get("category")
        tags = request.POST.get("tags")
        cancellation_policy = request.POST.get("cancellation_policy")
        # booking_link = request.POST.get("booking_link")
        status = request.POST.get("status")
        feed= request.POST.get("feed")

        # Create a new TravelPackage object with the form data
        package = TravelPackage(
            package_name=package_name,
            description=description,
            destination=destination,
            duration=duration,
            price=price,
            start_date=start_date,
            end_date=end_date,
            accommodation=accommodation,
            # meals=meals,
            transportation=transportation,
            activities=activities,
            inclusions=inclusions,
            exclusions=exclusions,
            # ratings=ratings,
            availability=availability,
            booking_deadline=booking_deadline,
            category=category,
            tags=tags,
            cancellation_policy=cancellation_policy,
            # booking_link=booking_link,
            status=status,
            feed=feed,
        )

        # Save the package to the database
        package.save()

        # Handle image uploads and save them
        for image in images:
            # Create and save an Image object for each uploaded image
            package_image = PackageImage(package=package, image=image)
            package_image.save()

        # Redirect or perform further actions as needed
        return redirect("/upload_package")  # Redirect to a package list page or any other page

    # This block handles the GET request, rendering the template for uploading a package
    return render(request, "add_package.html")



from datetime import date, timezone

@never_cache
@login_required(login_url='log')
def view_travel_packages(request):
    # Query the database to get a list of TravelPackage instances
    all_packages = TravelPackage.objects.all()
    search_date = request.GET.get('search_date')
    
    running_packages = all_packages.filter(feed='Post')
    paused_packages = all_packages.filter(Q(status='Pending') | Q(status='Running'), feed='Save')

    if search_date:
        try:
            search_date = date.fromisoformat(search_date)
            running_packages = running_packages.filter(start_date=search_date)
            paused_packages = paused_packages.filter(start_date=search_date)
        except ValueError:
            # Handle the case when an invalid date is entered
            pass

        # Add pagination
    # Add pagination
    paginator_running = Paginator(running_packages, 10)
    page_running = request.GET.get('page_running')
    running_packages = paginator_running.get_page(page_running)

    paginator_paused = Paginator(paused_packages, 10)
    page_paused = request.GET.get('page_paused')
    paused_packages = paginator_paused.get_page(page_paused)

    # Render the template with the queried data and pagination context
    return render(request, 'view_package.html', {'running_packages': running_packages, 'paused_packages': paused_packages})



from django.shortcuts import render, redirect, get_object_or_404
@never_cache
@login_required(login_url='log')
def edit_package(request, package_id):
    # Get the TravelPackage object based on the package_id
    package = get_object_or_404(TravelPackage, pk=package_id)

    if request.method == "POST":
        # Update the TravelPackage attributes based on the form data
        package.package_name = request.POST.get('package_name')
        package.description = request.POST.get('description')
        package.destination = request.POST.get('destination')
        package.duration = int(request.POST.get('duration'))
        package.price = float(request.POST.get('price'))
        package.start_date = request.POST.get('start_date')
        package.end_date = request.POST.get('end_date')
        package.accommodation = request.POST.get('accommodation')
        package.transportation = request.POST.get('transportation')
        package.activities = request.POST.get('activities')
        package.category = request.POST.get('category')
        package.tags = request.POST.get('tags')
        package.availability = int(request.POST.get('availability'))
        package.booking_deadline = request.POST.get('booking_deadline')
        package.cancellation_policy = request.POST.get('cancellation_policy')
        package.status = request.POST.get('status')
        package.feed = request.POST.get('feed')
        
        selected_inclusions = [request.POST.get(key) for key in request.POST if key.startswith("inclusions")]
        inclusions = ', '.join(selected_inclusions)
        package.inclusions = inclusions

        # Process and build a comma-separated string for exclusions
        selected_exclusions = [request.POST.get(key) for key in request.POST if key.startswith("exclusions")]
        exclusions = ', '.join(selected_exclusions)
        package.exclusions = exclusions
        # Save the updated TravelPackage
        package.save()

        return redirect('view_packages')  # Redirect to the view_packages page after editing

    # Render the edit_package template with the package data
    return render(request, 'edit_package.html', {'package': package})


def package_detail(request, package_id):
    package = get_object_or_404(TravelPackage, pk=package_id)
    from_upcoming_journeys = request.GET.get('from_upcoming_journeys')
    user= request.user
    passenger = Passenger.objects.filter(package=package, user=user).first()

    if from_upcoming_journeys:
        # If the 'from_upcoming_journeys' query parameter is present, hide the book button
        return render(request, 'package_detail.html', {'package': package, 'from_upcoming_journeys': from_upcoming_journeys})

    if request.user.is_authenticated:
        if hasattr(request.user, 'traveller'):
            profile = request.user.traveller  # Get or create the Traveller model instance
            return render(request, 'package_detail.html', {'profile': profile,'package': package,'from_upcoming_journeys': from_upcoming_journeys,'passenger':passenger})
        if request.user.is_traveller:
            profile = request.user.is_traveller  # Corrected from 'is_traveller' to 'traveller'
            return render(request, 'package_detail.html', {'profile': profile,'package': package,'from_upcoming_journeys': from_upcoming_journeys,'passenger':passenger})
    else:
        return render(request, 'package_detail.html', {'package': package,'from_upcoming_journeys': from_upcoming_journeys,'passenger':passenger})


from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from .models import Passenger
@never_cache
@login_required(login_url='log')
def add_passenger(request, package_id):
    package = get_object_or_404(TravelPackage, id=package_id)
    if request.method == 'POST':
        user_id = request.user.id
        package_id = package_id
        passenger_limit = request.POST.get('passenger-limit')
        passenger_name_list = request.POST.getlist('passenger_name')
        passenger_age_list = request.POST.getlist('passenger_age')
        proof_of_id_list = request.FILES.getlist('proof_of_id')
        children=request.POST.get('children')
        print(passenger_name_list,passenger_limit)
            
            # Check if the number of entries in all lists match
        if len(passenger_name_list) == len(passenger_age_list) == len(proof_of_id_list):
            total_passengers = len(passenger_name_list)

            flag = 0
            if total_passengers <= package.availability:

                for i in range(total_passengers):
                    passenger_name = passenger_name_list[i]
                    passenger_age = passenger_age_list[i]
                    proof_of_id = proof_of_id_list[i]

                    passenger = Passenger(
                        user_id=user_id,
                        package_id=package_id,
                        passenger_name=passenger_name,
                        passenger_age=passenger_age,
                        proof_of_id=proof_of_id
                    )
                    passenger.save()
                    flag = 1

                if flag == 0:
                    return render(request, 'passenger_info.html', {'package':package,'package_id': package_id})
                else:
                                    # Update availability
                    package.availability -= total_passengers
                    package.save()
                    booking = Booking(
                    package_id=package_id,
                    user_id=user_id,
                    status='Pending' , # Set the status to 'Confirmed' or 'Pending' as needed
                    passenger_limit=passenger_limit,
                    children=children
                    )
                    booking.save()
                    messages.success(request, "Your Booking Procedures have been Initialized. Stay Connected for getting further Updates.")

                    return redirect('/thome')
            else:
                messages.error(request, "Not enough availability for the selected package.")

        else:
            # Handle the case where the number of entries in lists don't match
            messages.success(request, "Upload ID Proof of Number of Passengers Entered for Booking.")
    return render(request, 'passenger_info.html', {'package':package,'package_id': package_id})
      

from django.shortcuts import render
from .models import TravelPackage
from django.utils import timezone
@never_cache
@login_required(login_url='log')
def upcoming_journeys(request):
    current_date = timezone.now().date()
    user = request.user

    # Retrieve upcoming journeys for the user
    upcoming_journeys = TravelPackage.objects.filter(start_date__gt=current_date, passenger__user=user).distinct()
    # Create a list to store confirmed bookings
    confirmed_bookings = Booking.objects.filter(user=user)

    # Create a list to store the package data with images and booking status
    packages_with_images = []

    for journey in upcoming_journeys:
        # Retrieve images related to the travel package
        package_images = PackageImage.objects.filter(package=journey)
        passengers = Passenger.objects.filter(user=user, package=journey)

        # Check if there is a confirmed booking for the current user
        booking = confirmed_bookings.filter(package=journey).first()

        # Append the travel package, its images, and booking status to the list
        packages_with_images.append({
            'package': journey,
            'images': package_images,
            'booking': booking,
            'passengers':passengers,
        })

    return render(request, 'upcoming_journeys.html', {'packages_with_images': packages_with_images})
@never_cache
@login_required(login_url='log')
def delete_passenger(request, passenger_id):
    passenger = get_object_or_404(Passenger, id=passenger_id)

    # Get the related booking
    booking = Booking.objects.get(package=passenger.package, user=request.user)

    # Delete the passenger
    passenger.delete()

    # Update passenger_limit in the Booking model
    booking.passenger_limit -= 1
    booking.save()

    # Update availability in the TravelPackage model
    passenger.package.availability += 1
    passenger.package.save()

    # Redirect to the passenger details page or any other appropriate page
    # You can change 'passenger_details' to the actual URL pattern name for your passenger details view
    return redirect('passenger_details', package_id=passenger.package.id)

@never_cache
@login_required(login_url='log')
def passenger_details(request, package_id):
    user=request.user
    package = get_object_or_404(TravelPackage, id=package_id)
    passengers = Passenger.objects.filter(package=package,user=user)

    return render(request, 'passenger_details.html', {'package': package, 'passengers': passengers})


@never_cache
@login_required(login_url='log')
def history_journeys(request):
    current_date = timezone.now().date()
    user = request.user  # Assuming you are using authentication

    # Retrieve historical journeys for the user
    history_journeys = TravelPackage.objects.filter(end_date__lt=current_date, passenger__user=user).distinct()
    packages_with_images = []

    for journey in history_journeys:
        # Retrieve images related to the travel package
        booking = Booking.objects.filter(
                user=user,
                package=journey,
                status='Confirmed'
            ).first()

            # Retrieve images related to the travel package
        package_images = PackageImage.objects.filter(package=booking.package)
        rating = Rating.objects.filter(user=user, package=booking.package).first()

            # Append the travel package, its images, and booking status to the list
        packages_with_images.append({
                'package': booking.package,
                'images': package_images,
                'booking': booking, 
                'rating': rating,
 # Include booking status in the dictionary
            })

    return render(request, 'history_journeys.html', {'packages_with_images': packages_with_images})

@never_cache
@login_required(login_url='log')
def ongoing_journeys(request):
    current_date = timezone.now().date()
    user = request.user  # Assuming you are using authentication

    # Retrieve historical journeys for the user
    ongoing_journeys = TravelPackage.objects.filter(start_date__lte=current_date ,end_date__gt=current_date, passenger__user=user).distinct()

    packages_with_images = []

    for journey in ongoing_journeys:
        # Check if the package is booked and confirmed by the user
        booking = Booking.objects.filter(
            user=user,
            package=journey,
            status='Confirmed'
        ).first()

        # Retrieve images related to the travel package
        package_images = PackageImage.objects.filter(package=booking.package)

        # Append the travel package, its images, and booking status to the list
        packages_with_images.append({
            'package': booking.package,
            'images': package_images,
            'booking': booking,  # Include booking status in the dictionary
        })

    return render(request, 'ongoing_journeys.html', {'packages_with_images': packages_with_images})


from django.db import transaction
from django.urls import reverse
@never_cache
@login_required(login_url='log')
def cancel_booking(request, package_id):
    # Retrieve the travel package and all passengers associated with the package and the logged-in user
    package = get_object_or_404(TravelPackage, pk=package_id)
    passengers = Passenger.objects.filter(package=package_id, user=request.user)

    # Check if there are passengers to cancel
    if passengers:
        try:
            # Store the length of passengers before deleting
            num_passengers = len(passengers)

            # Delete all passenger records
            passengers.delete()

            booking = Booking.objects.filter(user=request.user, package=package)
            booking.delete()

            # Update the availability of the travel package
            print(f"Before: {package.availability}")
            package.availability += num_passengers
            package.save()
            print(f"After: {package.availability}")

            messages.success(request, "Your bookings have been canceled successfully")

        except Exception as e:
            # Handle any exceptions that might occur during the database operations
            messages.error(request, f"An error occurred: {e}")

    else:
        # Handle the case where no bookings were found
        messages.error(request, "No bookings found for cancellation")

    return redirect(reverse('thome'))  # Redirect to 'thome' using the named URL pattern






from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Subquery, OuterRef
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from django.db.models import OuterRef, Subquery
from django.db.models import Count  # Add this import at the top of your views.py
from django.db.models import Sum

@staff_member_required
@never_cache
@login_required(login_url='log')
def upcoming_bookings(request):
    current_date = timezone.now().date()
    tour=TravelPackage.objects.all()
    
    for packages in tour:
        if current_date > packages.start_date:
            packages.status = 'Pending'
            packages.save()

    # Get the number of pending passengers for each package
    pending_passengers = Booking.objects.filter(package__in=tour, status='Pending') 
    pending_passengers_count= pending_passengers.aggregate(total_passenger_limit=Sum('passenger_limit'))['total_passenger_limit']


    running_packages = TravelPackage.objects.filter(
        status='Running',
        start_date__gt=current_date,feed='Post'
    )

    return render(request, 'upcoming_bookings.html', {'running_packages':running_packages,'package_pending_passengers': pending_passengers_count})



@staff_member_required
@never_cache
@login_required(login_url='log')
def verified_bookings(request):
    current_date = timezone.now().date()
    tour=TravelPackage.objects.all()
    
    for packages in tour:
        if current_date > packages.start_date:
            packages.status = 'Pending'
            packages.save()

    # Get the number of pending passengers for each package
    verified_passengers = Booking.objects.filter(package__in=tour, status='Confirmed') 
    pending_passengers_count= verified_passengers.aggregate(total_passenger_limit=Sum('passenger_limit'))['total_passenger_limit']


    running_packages = TravelPackage.objects.filter(
        status='Running',
        start_date__gt=current_date,feed='Post'
    )

    return render(request, 'verified_bookings.html', {'running_packages':running_packages,'package_pending_passengers': pending_passengers_count})



@login_required(login_url='log')
def package_requests(request, package_id):
    # Retrieve the package or return a 404 response if not found
    package = get_object_or_404(TravelPackage, id=package_id)

    # Retrieve booking information for the specific package with status 'Pending'
    bookings = Booking.objects.filter(package=package, status='Pending')
    
    # Create a set to store unique user IDs
    unique_users = set()

    # Create a list to store the relevant information for each unique user
    booking_details = []
    for booking in bookings:
        user_id = booking.user.id

        # Check if the user has already been added
        if user_id not in unique_users:
            user_info = {
                'package_id': booking.package.id,
                'location': booking.user.traveller.country,
                'package_name': booking.package.package_name,
                'start_date': booking.package.start_date,
                'user_email': booking.user.email,
                'passenger_details_url': reverse('passenger_count', args=[booking.package.id]),
                'verification_status': booking.status,
                'user_id': user_id,  # Add user id for the form action URL
            }
            booking_details.append(user_info)

            # Add user ID to the set
            unique_users.add(user_id)

    items_per_page = 10
    paginator = Paginator(booking_details, items_per_page)

    page = request.GET.get('page')
    try:
        bookings = paginator.page(page)
    except PageNotAnInteger:
        # If the page is not an integer, show the first page
        bookings = paginator.page(1)
    except EmptyPage:
        # If the page is out of range, show the last page
        bookings = paginator.page(paginator.num_pages)

    # Render the package_requests.html template with the collected information
    return render(request, 'package_requests.html', {'package': package, 'bookings': bookings})

@login_required(login_url='log')
def booking_list(request, package_id):
    # Retrieve the package or return a 404 response if not found
    package = get_object_or_404(TravelPackage, id=package_id)
    passengers = Passenger.objects.filter(package=package,status='Confirmed')
    verified_passengers = Booking.objects.filter( status='Confirmed',package=package) 
    sum_of_children = verified_passengers.aggregate(Sum('children'))['children__sum']
    print(sum_of_children)
    # Retrieve booking information for the specific package with status 'Pending'
    bookings = Booking.objects.filter(package=package, status='Confirmed')
    
    # Create a set to store unique user IDs
    unique_users = set()

    # Create a list to store the relevant information for each unique user
    booking_details = []
    for booking in bookings:
        user_id = booking.user.id

        # Check if the user has already been added
        if user_id not in unique_users:
            user_info = {
                'package_id': booking.package.id,
                'location': booking.user.traveller.country,
                'package_name': booking.package.package_name,
                'start_date': booking.package.start_date,
                'user_email': booking.user.email,
                'passenger_details_url': reverse('passenger_count', args=[booking.package.id]),
                'verification_status': booking.status,
                'user_id': user_id,  # Add user id for the form action URL
            }
            booking_details.append(user_info)

            # Add user ID to the set
            unique_users.add(user_id)

    items_per_page = 10
    paginator = Paginator(passengers, items_per_page)

    page = request.GET.get('page')
    try:
        passengers = paginator.page(page)
    except PageNotAnInteger:
        # If the page is not an integer, show the first page
        passengers = paginator.page(1)
    except EmptyPage:
        # If the page is out of range, show the last page
        passengers = paginator.page(paginator.num_pages)

    return render(request, 'passenger_count.html', {'package': package, 'passengers': passengers,'sum_of_children':sum_of_children})




from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

@never_cache
@login_required(login_url='log')
def passenger_count(request, package_id):
    # Retrieve the specific travel package
    package = get_object_or_404(TravelPackage, pk=package_id)

    # Retrieve all passengers for the selected package
    passengers = Passenger.objects.filter(package=package)


    # Set the number of items per page
    items_per_page = 10
    paginator = Paginator(passengers, items_per_page)

    page = request.GET.get('page')
    try:
        passengers = paginator.page(page)
    except PageNotAnInteger:
        # If the page is not an integer, show the first page
        passengers = paginator.page(1)
    except EmptyPage:
        # If the page is out of range, show the last page
        passengers = paginator.page(paginator.num_pages)

    return render(request, 'passenger_count.html', {'package': package, 'passengers': passengers})


from datetime import timedelta
@staff_member_required
@never_cache
@login_required(login_url='log')
def ongoing_bookings(request):
    current_date = timezone.now().date()
    tour=TravelPackage.objects.all()
    
    for packages in tour:
        if current_date == packages.start_date:
            packages.status = 'Pending'
            packages.save()


            end_date_condition = packages.start_date + timedelta(days=packages.duration) # Calculate the expected end date        # if current_date > end_date_condition:
        #     packages.status = 'Expired'
        #     packages.save()
            print(end_date_condition)
    # Get the number of pending passengers for each package
    pending_passengers = Booking.objects.filter(package__in=tour, status='Pending') 
    pending_passengers_count= pending_passengers.aggregate(total_passenger_limit=Sum('passenger_limit'))['total_passenger_limit']
    
    running_packages = TravelPackage.objects.filter(
        status='Pending',
        start_date__lte=current_date,
    )


    return render(request, 'ongoing_bookings.html', {'running_packages':running_packages,'package_pending_passengers': pending_passengers_count})



@login_required(login_url='log')
def ongoing_passengers(request, package_id):
    # Retrieve the package or return a 404 response if not found
    package = get_object_or_404(TravelPackage, id=package_id)

    # Retrieve booking information for the specific package with status 'Pending'
    bookings = Booking.objects.filter(package=package, status='Confirmed')
    
    # Create a set to store unique user IDs
    unique_users = set()

    # Create a list to store the relevant information for each unique user
    booking_details = []
    for booking in bookings:
        user_id = booking.user.id

        # Check if the user has already been added
        if user_id not in unique_users:
            user_info = {
                'package_id': booking.package.id,
                'location': booking.user.traveller.country,
                'package_name': booking.package.package_name,
                'start_date': booking.package.start_date,
                'user_email': booking.user.email,
                'passenger_details_url': reverse('passenger_count', args=[booking.package.id]),
                'verification_status': booking.status,
                'user_id': user_id,  # Add user id for the form action URL
            }
            booking_details.append(user_info)

            # Add user ID to the set
            unique_users.add(user_id)

    items_per_page = 10
    paginator = Paginator(booking_details, items_per_page)

    page = request.GET.get('page')
    try:
        bookings = paginator.page(page)
    except PageNotAnInteger:
        # If the page is not an integer, show the first page
        bookings = paginator.page(1)
    except EmptyPage:
        # If the page is out of range, show the last page
        bookings = paginator.page(paginator.num_pages)

    # Render the package_requests.html template with the collected information
    return render(request, 'ongoing_package_users.html', {'package': package, 'bookings': bookings})


@staff_member_required
@never_cache
@login_required(login_url='log')
def history_bookings(request):
    current_date = timezone.now().date()
    tour=TravelPackage.objects.all()
    
    for packages in tour:
        if current_date > packages.end_date:
            packages.status = 'Pending'
            packages.save()

    # Get the number of pending passengers for each package
    pending_passengers = Booking.objects.filter(package__in=tour, status='Pending') 
    pending_passengers_count= pending_passengers.aggregate(total_passenger_limit=Sum('passenger_limit'))['total_passenger_limit']


    running_packages = TravelPackage.objects.filter(
        status='Pending',
        end_date__lt=current_date,
    )

    return render(request, 'history_bookings.html', {'running_packages':running_packages,'package_pending_passengers': pending_passengers_count})



@login_required(login_url='log')
def history_passengers(request, package_id):
    # Retrieve the package or return a 404 response if not found
    package = get_object_or_404(TravelPackage, id=package_id)

    # Retrieve booking information for the specific package with status 'Pending'
    bookings = Booking.objects.filter(package=package, status='Confirmed')
    
    # Create a set to store unique user IDs
    unique_users = set()

    # Create a list to store the relevant information for each unique user
    booking_details = []
    for booking in bookings:
        user_id = booking.user.id
        ratings= Rating.objects.filter(package=package,user=booking.user)
        # Check if the user has already been added
        if user_id not in unique_users:
            user_info = {
                'package_id': booking.package.id,
                'location': booking.user.traveller.country,
                'package_name': booking.package.package_name,
                'start_date': booking.package.start_date,
                'user_email': booking.user.email,
                'passenger_details_url': reverse('passenger_count', args=[booking.package.id]),
                'verification_status': booking.status,
                'user_id': user_id,  # Add user id for the form action URL
                'stars': ratings.first().stars if ratings.exists() else 0,
            }
            booking_details.append(user_info)

            # Add user ID to the set
            unique_users.add(user_id)

    items_per_page = 10
    paginator = Paginator(booking_details, items_per_page)

    page = request.GET.get('page')
    try:
        bookings = paginator.page(page)
    except PageNotAnInteger:
        # If the page is not an integer, show the first page
        bookings = paginator.page(1)
    except EmptyPage:
        # If the page is out of range, show the last page
        bookings = paginator.page(paginator.num_pages)

    # Render the package_requests.html template with the collected information
    return render(request, 'history_package_users.html', {'package': package, 'bookings': bookings})


@never_cache
@login_required(login_url='log')
def update_booking_status(request, user_id, package_id):
    if request.method == 'POST':
        status = request.POST.get('status')
        package = TravelPackage.objects.get(pk=package_id)

        if status == 'Pending':
            subject = f'Booking Status Update for Booking ID {package.package_name}'
            message = f'Your booking status has been updated to: Pending'
        elif status == 'Confirmed':
            subject = f'Booking Status Update for Booking ID {package.package_name}'
            login_url = request.build_absolute_uri(reverse('upcoming_journeys'))  # Adjust the URL name if needed
            message = f"Your booking status has been Confirmed. Now you can Login to NIWI TRAVELS and make the payment. Click here to '{login_url}'"
        elif status == 'Cancelled':
            subject = f'Booking Status Update for Booking ID {package.package_name}'
            login_url = request.build_absolute_uri(reverse('log'))  # Adjust the URL name if needed
            message = f"Your booking status has been Cancelled due to the incorrect entry of information of the Passengers. Login to NIWI TRAVELS and reenter the data again. We are happy to help you. Click here to login '{login_url}'"
        else:
            # Handle an invalid status here (if needed)
            return render(request, 'invalid_status.html')  # Create this template

        # Use get() instead of filter() to get a single Booking instance
        booking = Booking.objects.filter(user_id=user_id, package_id=package_id).first()

        if booking:
            # Update the booking status
            booking.status = status
            booking.save()

            # Send an email notification
            from_email = 'godwinbmenachery2024a@mca.ajce.in'  # Replace with your email
            recipient_list = [booking.user.email]  # Use the email of the booking user

            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            return redirect(upcoming_bookings)
        else:
            # Handle the case where there is no matching booking
            return render(request, 'booking_not_found.html')  # Create this template

from decimal import Decimal
@never_cache
@login_required(login_url='log')
def payment(request, booking_id):
    booking = get_object_or_404(Booking, pk=booking_id)
    payment = Payment.objects.filter(booking=booking).first()

    # Get the TravelPackage instance associated with the booking
    travel_package = booking.package
    total_amount = travel_package.price * booking.passenger_limit

    # Calculate GST amount
    gst_rate = Decimal('0.15')  # 15% GST rate
    gst_amount = total_amount * gst_rate

    # Calculate final price including GST
    final_price = total_amount + gst_amount

    # You can include any payment processing logic here if needed

    return render(request, 'payment.html', {'booking': booking, 'travel_package': travel_package, 'total_amount': total_amount, 'gst_amount': gst_amount, 'final_price': final_price, 'payment': payment})



import razorpay
from django.views.decorators.csrf import csrf_exempt
def pay(request,booking_id):
    flag=0
    
    if request.method == 'POST':
        book = get_object_or_404(Booking, pk=booking_id)
        client = razorpay.Client(auth=("rzp_test_PvsGkN41iQ2AJL", "6RXQIZzmrKaWRPW2tUuBbiP6"))
        travel_package = book.package
        total_amount = travel_package.price * book.passenger_limit
        gst_rate = Decimal('0.15')  # 15% GST rate
        gst_amount = total_amount * gst_rate

        # Calculate final price including GST
        final_price = total_amount + gst_amount
        book_amount = int(final_price* 100)  # Example amount
        data = {
            "amount": book_amount,
            "currency": "INR",
            "receipt": f"order_rcptid_{booking_id}"  # Use order ID to generate a unique receipt ID
        }
        payment = client.order.create(data=data)
        if payment:
            # Save the payment details to your database or perform any necessary action

            # Render the pay.html page with payment details
            return render(request, 'pay.html', {'payment': payment, 'order': book})

    return HttpResponse("Invalid request")

def success(request,booking_id):
    book = get_object_or_404(Booking, pk=booking_id)
    travel_package = book.package
    total_amount = travel_package.price * book.passenger_limit
    book_amount = int(total_amount* 100)  # Example amount

    client = razorpay.Client(auth=("rzp_test_PvsGkN41iQ2AJL", "6RXQIZzmrKaWRPW2tUuBbiP6"))
    data = {
            "amount": book_amount,
            "currency": "INR",
            "receipt": f"order_rcptid_{booking_id}"  # Use order ID to generate a unique receipt ID
        }
    payment = client.order.create(data=data)
    new_payment = Payment(
            booking=book,
            razor_pay_order_id=payment['id'],
            amount=book_amount,
            is_paid=True,
            user=request.user  # Assuming the user is authenticated and initiating the payment
        )
    new_payment.save()
    return redirect('/thome')


from django.shortcuts import get_object_or_404, render

# Assuming you have a method to get the 'Pending' Booking instance

    

# Now, use this function as the default for the ForeignKey
def get_default_booking_instance():
    try:
        return Booking.objects.get(status='Pending')
    except Booking.DoesNotExist:
        return None
    
import openpyxl
def export_passenger_data_to_excel(request,package_id):
    # Create a new Excel workbook and add a worksheet
    package = get_object_or_404(TravelPackage, pk=package_id)

    passengers = Passenger.objects.filter(package=package)  # This is just an example, adjust based on your actual model and query logic

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['Name', 'Age', 'Proof of ID']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add passenger data to the worksheet
    for row_num, passenger in enumerate(passengers, 2):
        worksheet.cell(row=row_num, column=1, value=passenger.passenger_name)
        worksheet.cell(row=row_num, column=2, value=passenger.passenger_age)
        worksheet.cell(row=row_num, column=3, value=passenger.proof_of_id.url)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=passenger_data.xlsx'
    workbook.save(response)

    return response

from django.http import JsonResponse
def submit_rating(request, package_id, stars):
    if request.method == 'POST':
        user = request.user
        package = TravelPackage.objects.get(pk=package_id)

        # Check if the form data is valid

        if not request.POST.get('stars') or not request.POST.get('description'):
            messages.error(request, 'Please select a rating and provide a description')
            return redirect(history_journeys)

        # Retrieve rating from the form data
        rating_value = request.POST.get('stars')

        # Check if the user has already rated the package
        try:
            rating = Rating.objects.get(user=user, package=package)
            rating.stars = rating_value
            rating.description = request.POST.get('description', '')
            rating.save()
        except Rating.DoesNotExist:
            rating = Rating.objects.create(user=user, package=package, stars=rating_value, description=request.POST.get('description', ''))

        return redirect(history_journeys)

    return redirect(history_journeys)

